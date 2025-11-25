from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class ExhibitorRow:
    index: int
    name: str
    zone: str
    date: str
    booth: int | None


def load_rows(xlsx_path: Path) -> list[ExhibitorRow]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    if header[1:5] != (
        "Name of Company/Organization",
        "Zone",
        "Date",
        "Booth No.",
    ):
        raise RuntimeError(f"Unexpected header: {header}")

    exhibitors: list[ExhibitorRow] = []
    for _excel_row_index, row in enumerate(rows[1:], start=2):
        if row[1] is None:
            continue
        idx, name, zone, date, booth = row[:5]
        exhibitors.append(
            ExhibitorRow(
                index=int(idx),
                name=str(name).strip(),
                zone=str(zone).strip(),
                date=str(date).strip(),
                booth=int(booth) if booth is not None else None,
            )
        )
    return exhibitors


def to_swift_identifier(name: str) -> str:
    # Simple, safe ID based on index in Excel; name will be included in data.
    return name


def generate_swift(exhibitors: list[ExhibitorRow]) -> str:
    lines: list[str] = []
    lines.append("// Generated from Content/HKBU Career Fair 2025 Exhibitor list.xlsx")
    lines.append("// Do not edit by hand – run scripts/export_exhibitors.py instead.\n")
    lines.append("import Foundation\n")
    lines.append("// All exhibitors from the official Excel file.")
    lines.append("let allExhibitors: [Exhibitor] = [")

    for row in exhibitors:
        safe_name = row.name.replace('"', '\\"')
        safe_zone = row.zone.replace('"', '\\"')
        safe_date = row.date.replace('"', '\\"')
        # For now, default all to recruitment (blue); can refine using Job Description files.
        purpose = " .recruitment"
        buhub_url = "https://buhub.hkbu.edu.hk/company/" + str(row.index)
        booth_literal = str(row.booth) if row.booth is not None else "nil"

        lines.append(
            f'    Exhibitor(id: "ex-{row.index}", '
            f'name: "{safe_name}", '
            f'industrySector: "{safe_zone}", '
            f'dateAttendance: "{safe_date}", '
            f'purpose: .recruitment, '
            f'buhubURL: URL(string: "{buhub_url}")!, '
            f'boothNumber: {booth_literal}),'
        )

    lines.append("]\n")
    return "\n".join(lines)


def main() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    xlsx_path = base_dir / "Content" / "HKBU Career Fair 2025 Exhibitor list.xlsx"

    exhibitors = load_rows(xlsx_path)
    swift_source = generate_swift(exhibitors)

    output_path = base_dir / "Models" / "ExhibitorData.swift"
    output_path.write_text(swift_source, encoding="utf-8")
    print(f"Wrote {len(exhibitors)} exhibitors to {output_path}")


if __name__ == "__main__":
    main()



